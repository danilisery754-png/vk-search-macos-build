from __future__ import annotations

import csv
import io
from dataclasses import asdict, dataclass
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass(frozen=True, slots=True)
class ResultExportRow:
    group_name: str
    url: str
    message_status: str
    suggested_status: str
    destination: str
    account_name: str
    reason: str


_HEADERS = ["Группа", "Ссылка", "ЛС", "Предложка", "Куда получилось", "Аккаунт", "Причина"]


def _clean(value: object) -> str:
    return str(value if value is not None else "").replace("\t", " ").replace("\r", " ").replace("\n", " ")


def _values(row: ResultExportRow) -> list[str]:
    return [_clean(value) for value in asdict(row).values()]


def links_text(rows: Iterable[ResultExportRow]) -> str:
    return "\n".join(_clean(row.url) for row in rows)


def rows_tsv(rows: Iterable[ResultExportRow]) -> str:
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(_HEADERS)
    writer.writerows(_values(row) for row in rows)
    return output.getvalue().rstrip("\n")


def rows_csv(rows: Iterable[ResultExportRow]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(_HEADERS)
    writer.writerows(_values(row) for row in rows)
    return output.getvalue().encode("utf-8-sig")


def rows_xlsx(rows: Iterable[ResultExportRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результаты"
    sheet.append(_HEADERS)
    for row in rows:
        sheet.append(_values(row))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    widths = (34, 42, 18, 18, 22, 24, 55)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
